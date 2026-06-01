#!/usr/bin/env python3
"""
build_workbook.py — siege.db + reference/ → data/competitions/{id}/workbook.xlsx

Usage:
  python3 build_workbook.py --comp 100
  python3 build_workbook.py --comp 101
"""
import argparse
import json
import os
from collections import defaultdict

import sqlite3

HERE = os.path.dirname(os.path.abspath(__file__))
REF  = os.path.join(HERE, "reference")
DATA = os.path.join(HERE, "data")


# ── Reference data ────────────────────────────────────────────────────────────

def _load_ref():
    maps = json.load(open(os.path.join(REF, "maps.json")))
    return (
        maps["map_order"],
        maps["map_sites"],
        maps["map_picks_layout"],
    )


def _load_config(comp_id):
    """Load competition metadata from competitions.json (single source of truth)."""
    catalog_path = os.path.join(DATA, "competitions.json")
    if os.path.exists(catalog_path):
        for entry in json.load(open(catalog_path)):
            if entry["id"] == int(comp_id):
                return entry
    raise FileNotFoundError(
        f"Competition {comp_id} not found in competitions.json. "
        f"Run: python3 run.py sync-catalog"
    )


# ── DB queries ────────────────────────────────────────────────────────────────

def _query(comp_id):
    db_path = os.path.join(DATA, "siege.db")
    conn = sqlite3.connect(db_path)

    picks = conn.execute(
        "SELECT match_id, date, map, round, side, operator, site FROM picks WHERE comp_id=?",
        (comp_id,)
    ).fetchall()

    bans = conn.execute(
        "SELECT match_id, map, side, operator, game_rounds FROM bans WHERE comp_id=?",
        (comp_id,)
    ).fetchall()

    rounds = conn.execute(
        "SELECT match_id, map, site, def_win FROM rounds WHERE comp_id=?",
        (comp_id,)
    ).fetchall()

    matches = conn.execute(
        "SELECT DISTINCT match_id FROM processed_matches WHERE comp_id=? ORDER BY match_id",
        (comp_id,)
    ).fetchall()

    conn.close()
    return picks, bans, rounds, [r[0] for r in matches]


# ── Aggregation ───────────────────────────────────────────────────────────────

def _aggregate(picks, bans, rounds):
    opstats = defaultdict(lambda: {"banned": 0, "value": 0, "sites": defaultdict(int)})
    for _, _date, map_name, _round, side, operator, site in picks:
        opstats[(map_name, side, operator)]["sites"][site] += 1
    for _, map_name, side, operator, game_rounds in bans:
        s = opstats[(map_name, side, operator)]
        s["banned"] += 1
        s["value"] += game_rounds
    site_rounds = defaultdict(lambda: {"rounds": 0, "wins": 0})
    for _, map_name, site, def_win in rounds:
        site_rounds[(map_name, site)]["rounds"] += 1
        site_rounds[(map_name, site)]["wins"] += def_win
    return opstats, site_rounds


def _games_per_map(picks, bans, rounds):
    games = defaultdict(set)
    for mid, map_name, site, _ in rounds:
        games[map_name].add(mid)
    return {m: len(mids) for m, mids in games.items()}


# ── Match name helpers ────────────────────────────────────────────────────────

def _matchup(slug):
    s = slug.replace("mjr-intl-", "").replace("lsr-", "").replace("lnr-", "")
    if "-vs-" in s:
        a, b = s.split("-vs-", 1)
        return f"{_team(a)} vs {_team(b)}"
    return s.replace("-", " ").title()


def _team(slug):
    return slug.replace("-", " ").title()


# ── Workbook construction ─────────────────────────────────────────────────────

def _styled(ws, row_idx, ncols, bold_font, reg_font, bold=False):
    for c in range(1, ncols + 1):
        ws.cell(row=row_idx, column=c).font = bold_font if bold else reg_font


def _deviations(mp_map, mp_site, map_picks_layout):
    dev = {}
    for m, sites in map_picks_layout:
        m_rounds, m_wr = mp_map.get(m, ("", ""))
        if not isinstance(m_rounds, (int, float)) or not m_rounds:
            continue
        for s in sites:
            s_rounds, s_wr = mp_site.get((m, s), ("", ""))
            if isinstance(s_rounds, (int, float)) and s_rounds:
                dev[(m, s)] = (s_wr - m_wr) * 100 * (s_rounds / m_rounds)
    return dev


def build_workbook(comp_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    comp_id = int(comp_id)
    map_order, map_sites, map_picks_layout = _load_ref()
    config = _load_config(comp_id)
    comp_name = config.get("name", f"Competition {comp_id}")
    atk_list  = config.get("atk_list", [])
    def_list  = config.get("def_list", [])

    picks, bans, rounds, match_ids = _query(comp_id)

    # Load match slugs from matches.json for Match URLs tab
    matches_path = os.path.join(DATA, "competitions", str(comp_id), "matches.json")
    match_slugs = json.load(open(matches_path)) if os.path.exists(matches_path) else {}

    # Maps played per match from rounds data
    maps_per_match = defaultdict(set)
    for mid, map_name, _site, _dw in rounds:
        maps_per_match[str(mid)].add(map_name)

    opstats, site_rounds = _aggregate(picks, bans, rounds)
    games_per_map = _games_per_map(picks, bans, rounds)

    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(name="Arial", bold=True)
    reg  = Font(name="Arial")

    # ── Match URLs ──
    ws = wb.create_sheet("Match URLs")
    ws.append(["Match ID", "Matchup", "Maps Played", "URL"])
    _styled(ws, 1, 4, bold, reg, True)
    for i, mid in enumerate(sorted(match_slugs, key=lambda x: int(x)), start=2):
        slug = match_slugs[mid]
        url  = f"https://siege.gg/matches/{mid}-{slug}"
        maps_played = ", ".join(sorted(maps_per_match.get(mid, [])))
        ws.append([int(mid), _matchup(slug), maps_played, url])
        _styled(ws, i, 4, bold, reg)
    for col, w in (("A", 10), ("B", 36), ("C", 40), ("D", 60)):
        ws.column_dimensions[col].width = w

    # ── Map Picks ──
    mp_map, mp_site = {}, {}
    for m, sites in map_picks_layout:
        m_rounds = sum(site_rounds[(m, s)]["rounds"] for s in sites)
        m_wins   = sum(site_rounds[(m, s)]["wins"]   for s in sites)
        mp_map[m] = (m_rounds or "", round(100 * m_wins / m_rounds) if m_rounds else "")
        for s in sites:
            sr = site_rounds[(m, s)]
            mp_site[(m, s)] = (
                sr["rounds"] or "",
                round(100 * sr["wins"] / sr["rounds"]) if sr["rounds"] else "",
            )
    dev = _deviations(mp_map, mp_site, map_picks_layout)
    _write_map_picks(wb, mp_map, mp_site, dev, atk_list, def_list, map_picks_layout, bold, reg)

    # ── Per-map sheets ──
    for m in map_order:
        sites = map_sites[m]
        total_rounds = sum(site_rounds[(m, s)]["rounds"] for s in sites)
        title = f"{m} | {comp_name} | {games_per_map.get(m, 0)} games | {total_rounds} rounds"
        side_data = {}
        for side in ("Defense", "Attack"):
            ops = [(op, v) for (mm, sd, op), v in opstats.items() if mm == m and sd == side]
            ops.sort(key=lambda kv: (-sum(kv[1]["sites"].values()), -kv[1]["value"], kv[0]))
            side_data[side] = [
                (op, v["banned"], v["value"], [v["sites"].get(s, 0) for s in sites])
                for op, v in ops
            ]
        _write_map_sheet(wb, m, title, sites, side_data, dev, bold, reg)

    # ── Raw Data ──
    _write_raw_data_sheet(wb, picks, match_slugs, bold, reg)

    # ── Sheet order ──
    _rank = {"Map Picks": 0, "Match URLs": 1, "Raw Data": 99}
    wb._sheets.sort(key=lambda s: _rank.get(s.title, 2))

    out_dir  = os.path.join(DATA, "competitions", str(comp_id))
    out_path = os.path.join(out_dir, "workbook.xlsx")
    os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return wb


def _write_map_picks(wb, mp_map, mp_site, dev, atk_list, def_list,
                     map_picks_layout, bold, reg):
    ws = wb.create_sheet("Map Picks")
    ws.cell(1, 2, "Rounds"); ws.cell(1, 3, "Win Rate")
    _styled(ws, 1, 3, bold, reg, True)
    ri = 2
    for m, sites in map_picks_layout:
        rounds, wr = mp_map.get(m, ("", ""))
        ws.cell(ri, 1, m).font = bold
        if rounds != "": ws.cell(ri, 2, rounds)
        if wr    != "": ws.cell(ri, 3, wr)
        ri += 1
        for s in sites:
            srounds, swr = mp_site.get((m, s), ("", ""))
            ws.cell(ri, 1, s)
            if srounds != "": ws.cell(ri, 2, srounds)
            if swr     != "": ws.cell(ri, 3, swr)
            if (m, s) in dev:
                ws.cell(ri, 4, round(dev[(m, s)], 8))
            _styled(ws, ri, 4, bold, reg)
            ri += 1
        ri += 1
    if atk_list:
        ws.cell(3, 7, "Atk").font = bold
        for i, op in enumerate(atk_list): ws.cell(4 + i, 7, op)
    if def_list:
        ws.cell(3, 9, "Def").font = bold
        for i, op in enumerate(def_list): ws.cell(4 + i, 9, op)
    for col, w in (("A", 24), ("B", 10), ("C", 10), ("D", 14), ("G", 14), ("I", 14)):
        ws.column_dimensions[col].width = w


def _write_map_sheet(wb, m, title, sites, side_data, dev, bold, reg):
    ws = wb.create_sheet(m[:31])
    n = len(sites)
    prio0 = 4 + n
    ws.cell(1, 1, title).font = bold
    r = 3
    for side in ("Defense", "Attack"):
        ws.cell(r, 1, side).font = bold
        for j, s in enumerate(sites):
            ws.cell(r, prio0 + j, s).font = bold
        r += 1
        ws.cell(r, 1, "Operator"); ws.cell(r, 2, "Times Banned")
        ws.cell(r, 3, "Value (rounds x bans)")
        for j, s in enumerate(sites):
            ws.cell(r, 4 + j, s)
            if (m, s) in dev:
                ws.cell(r, prio0 + j, round(dev[(m, s)], 8))
        _styled(ws, r, prio0 + n - 1, bold, reg, True)
        r += 1
        for op, bans, value, apps in side_data[side]:
            ws.cell(r, 1, op); ws.cell(r, 2, bans); ws.cell(r, 3, value)
            for j in range(n):
                ws.cell(r, 4 + j, apps[j])
                ws.cell(r, prio0 + j, value + apps[j])
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 16
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 14


def _write_raw_data_sheet(wb, picks, match_slugs, bold, reg):
    ws = wb.create_sheet("Raw Data")
    headers = ["Match Name", "Date", "Map", "Round", "Side", "Operator", "Scenario", "Times"]
    ws.append(headers)
    _styled(ws, 1, len(headers), bold, reg, True)
    for mid, date, map_name, round_num, side, operator, site in sorted(
        picks, key=lambda x: (x[0], x[2], x[3])
    ):
        slug = match_slugs.get(str(mid), "")
        match_name = _matchup(slug) if slug else str(mid)
        ws.append([match_name, date, map_name, round_num, side, operator, site, 1])
    for col, w in (("A", 36), ("B", 12), ("C", 18), ("D", 8),
                   ("E", 10), ("F", 16), ("G", 26), ("H", 8)):
        ws.column_dimensions[col].width = w


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--comp", type=int, required=True, help="Competition ID")
    args = ap.parse_args()
    wb = build_workbook(args.comp)
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
